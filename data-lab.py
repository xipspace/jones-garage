import json
import os
import re
import openpyxl

# file handler

def select_item(items, item_type):
    print(f"{item_type}s:")
    for i, item in enumerate(items, 1):
        print(f"{i}. {item}")
    selected_index = input(f"Select the number of the {item_type} to operate: ")
    try:
        selected_index = int(selected_index) - 1
        if 0 <= selected_index < len(items):
            return items[selected_index]
        else:
            print(f"Invalid {item_type} selection.")
            return None
    except ValueError:
        print("Invalid input. Please enter a number.")
        return None

def list_files(extension):
    files = [file for file in os.listdir() if file.endswith(extension)]
    if files:
        return select_item(files, 'file')
    else:
        print("No files found with the specified extension.")
        return None

# Function to open the workbook
def open_workbook(file_name):
    try:
        workbook = openpyxl.load_workbook(file_name, read_only=True)
        return workbook
    except Exception as e:
        print(f"Error occurred while opening the workbook: {e}")
        return None

# Updated list_tabs function to accept a workbook object
def list_tabs(workbook):
    try:
        tabs = workbook.sheetnames
        return select_item(tabs, 'tab')
    except Exception as e:
        print(f"Error occurred while listing tabs: {e}")
        return None

# Updated list_headers function to accept a workbook object
def list_headers(workbook, sheet_name):
    try:
        sheet = workbook[sheet_name]
        headers = [cell.value for cell in sheet[1]]
        print("Headers:")
        for i, header in enumerate(headers, 1):
            print(f"{i}. {header}")
        selected_index = input("Select the number of the header to operate: ")
        try:
            selected_index = int(selected_index) - 1
            if 0 <= selected_index < len(headers):
                return headers[selected_index]
            else:
                print("Invalid header selection.")
                return None
        except ValueError:
            print("Invalid input. Please enter a number.")
            return None
    except Exception as e:
        print(f"Error occurred while reading headers: {e}")
        return None

# model


def treat_code(code):
    treated_code = re.sub(r'[^a-zA-Z0-9]', '', code).lower()
    treated_code = treated_code.replace('pdf', '')
    return treated_code

def extract_data(workbook, sheet_name):
    try:
        json_data = {}
        
        category = input("Category for the Object: ")
        
        # Assuming sheet_name is the name of the sheet, we need to retrieve the actual sheet object
        sheet = workbook[sheet_name]
        
        # Retrieve the headers from the first row of the sheet
        header_row = [cell.value.lower() for cell in sheet[1] if cell.value]
        
        # Find the indices of 'code' and 'label' headers
        code_index, label_index = header_row.index('code'), header_row.index('label')
        
        # Iterate over rows starting from the second row
        counter = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Extract code and label values
            code, label = row[code_index], row[label_index]
            
            if code and label is not None:
                uid = treat_code(code)
                if category:
                    json_content = {
                    "code": code,
                    "category": category,
                    "label": label,
                    "properties": { header_row[i]: row[i] for i in range(len(row)) if i not in {code_index, label_index} and row[i] is not None },
                    "restrictions": {}
                }
                else:
                    json_content = {
                    "code": code,
                    # "category": category,
                    "label": label,
                    "properties": { header_row[i]: row[i] for i in range(len(row)) if i not in {code_index, label_index} and row[i] is not None },
                    "restrictions": {}
                }
                # print(uid)
                json_data[uid] = json_content
                counter += 1
        
        print(f"{counter} objects created")
        if json_data:
            json_file = write_json(json_data, sheet_name)
          
        return json_data, json_file

    except Exception as e:
        print(f"Error occurred while extracting data: {e}")
        return None, None



def extract_cpl(json_data, json_file, workbook):
    try:

        sheet_name = list_tabs(workbook)
        sheet = workbook[sheet_name]

        # Get the index of the column where codes are stored
        code_column_index = int(input("Enter the number of the column where codes are stored: ")) - 1
        value_column_index = int(input("Enter the number of the column for values: ")) - 1

        # Extract and treat codes
        cpl_object = {}
        for row in sheet.iter_rows(min_row=2, min_col=code_column_index + 1, max_col=value_column_index + 1, values_only=True):
            code_value = row[0]
            cpl_value = row[value_column_index]
            if code_value:
                treated_code = treat_code(str(code_value))
                cpl_object[treated_code] = cpl_value.lower()

        # Compare extracted CPL data with JSON data
        if json_data:
            changes_count = 0
            for code, value in cpl_object.items():
                if code in json_data:
                    if cpl_object[code] == "acceptable":
                        json_data[code]["restrictions"]["cpl"] = False
                        changes_count += 1
                    else:
                        json_data[code]["restrictions"]["cpl"] = True
                        changes_count += 1

        print(f"Changes made: {changes_count}")

        if json_data and json_file:
            with open(json_file, 'w') as json_file:
                json.dump(json_data, json_file, indent=2)
                print(f"Updated JSON data written to {json_file}")

    except Exception as e:
        print(f"Error occurred while extracting and processing data: {e}")



# routines for JSON


# merge ops

def json_merge_ops(data, filename):
    
    try:
        counter = 0
        
        for key, value in data.items():
            # print(key)
            if "properties" in value:
                counter += 1
                # print(value)
                properties = value["properties"]
                # print(properties)
                ongoing = properties.get("ongoing", False)
                completed = properties.get("completed", False)
                operations = "none"
                
                if ongoing and completed:
                    operations = "both"
                elif ongoing:
                    operations = "ongoing"
                elif completed:
                    operations = "completed"

                # print(operations)
                
                updated_properties = {}
                for up_key, up_value in properties.items():
                    if up_key == "ongoing":
                        updated_properties["operations"] = operations
                    updated_properties[up_key] = up_value
                
                value["properties"] = updated_properties
                
                # print(updated_properties)
        
        print(f"{counter} objects updated")

        # clean the old keys
        remove_ongoing_completed_keys(data)
        
        with open(filename, 'w') as filename:
            json.dump(data, filename, indent=2)
        
        return data
        
    except Exception as e:
        print(f"Error occurred while merging JSON file: {e}")
        return None


# remove ongoing completed


def remove_ongoing_completed_keys(data):
    for key, value in data.items():
        if 'properties' in value:
            properties = value['properties']
            if 'ongoing' in properties:
                del properties['ongoing']
            if 'completed' in properties:
                del properties['completed']




# create aggregate

def json_update_aggregate(data, filename):
    
    try:
        counter = 0
        
        for key, value in data.items():
            # print(key)
            if "label" in value:
                label = value['label'].lower()
                project = "project" in label
                location = "location" in label
                
                properties = value.get("properties", {})
                
                if project and location:
                    properties["agg"] = "both"
                elif project:
                    properties["agg"] = "project"
                elif location:
                    properties["agg"] = "location"
                else:
                    properties["agg"] = "none"
                
                value['properties'] = properties
                counter += 1
        
        print(f"{counter} objects updated")

        with open(filename, 'w') as filename:
            json.dump(data, filename, indent=2)
        
        return data
        
    except Exception as e:
        print(f"Error occurred while updating JSON file: {e}")
        return None
    

# fix blanket AI

def json_update_blanket(data, filename):
    try:
        counter = 0
        
        for key, value in data.items():
            label = value.get("label", "").lower()
            if ('blanket additional' in label or 'blanket ai' in label) and value.get("properties"):
                properties = value["properties"]
                if "blanket" in properties and not properties["blanket"]:
                    properties["blanket"] = True
                    counter += 1
        
        print(f"{counter} objects updated")
        
        with open(filename, 'w') as filename:
            json.dump(data, filename, indent=2)
        
        return data
        
    except Exception as e:
        print(f"Error occurred while updating JSON file: {e}")
        return None


# fix blanket

# extract x days notice of cancellation
# read label
# turn noc into string

# remove none values





# json operations


def write_json(json_data, sheet_name):
    try:
        # Construct the file name
        file_name = f"index_{sheet_name.lower().replace(' ', '_')}.json"
        
        with open(file_name, "w") as json_file:
            json.dump(json_data, json_file, indent=2)

        print(f"JSON data has been written to {file_name}")
        
        return file_name

    except Exception as e:
        print(f"Error occurred while writing JSON file: {e}")
        return None



def load_json(file_name):
    try:
        with open(file_name, 'r') as json_file:
            json_data = json.load(json_file)
        return json_data

    except Exception as e:
        print(f"Error occurred while loading JSON file: {e}")
        return None




def show_menu():
    print("Menu:")
    print("1. Excel Selector")
    print("2. Extract Index to JSON")
    print("3. Extract CPL to Index")
    print("4. JSON Selector")
    print("5. Fix Operations & Aggregate")
    print("6. Fix Blanket AI")
    print("0. Quit")

def main():
    excel_file = None
    json_file = None
    workbook = None

    while True:
        show_menu()
        choice = input("Enter your choice: ")
        if choice == "1":
            excel_file = list_files('xlsx')
            if excel_file:
                workbook = open_workbook(excel_file)
                if workbook:
                    excel_sheet = list_tabs(workbook)
                else:
                    print("No Sheet data.")
            else:
                print("No Excel data.")

            # excel_index = list_headers(workbook, excel_sheet)
        elif choice == "2":
            if workbook and excel_sheet:
                json_data, json_file = extract_data(workbook, excel_sheet)
            else:
                print("Missing references.")
        elif choice == "3":
            if json_data and json_file and workbook:
                extract_cpl(json_data, json_file, workbook)
            else:
                print("Missing references.")
            # json_data = extract_cpl(json_data, json_file, workbook)
        elif choice == "4":
            json_file = list_files('json')
            if json_file:
                json_data = load_json(json_file)
            else:
                print("No JSON data.")
        elif choice == "5":
            if json_data and json_file:
                json_data = json_merge_ops(json_data, json_file)
                json_data = json_update_aggregate(json_data, json_file)
            else:
                print("No JSON data")
        elif choice == "6":
            if json_data and json_file:
                json_data = json_update_blanket(json_data, json_file)
            else:
                print("No JSON data")
        elif choice == "0":
            print("Exiting...")
            if workbook:
                workbook.close()
            break
        else:
            print("Invalid choice. Please enter a valid option.")

if __name__ == "__main__":
    main()