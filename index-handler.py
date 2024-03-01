import openpyxl
import json
import os
import re

class ExcelDataExtractor:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(self.file_path)
        self.uids = []

    def get_selected_sheet(self, tab_choice):
        return self.workbook[self.workbook.sheetnames[tab_choice]]

    def count_non_empty_rows_and_columns(self, sheet):
        non_empty_rows = sum(1 for row in sheet.rows if any(cell.value is not None for cell in row))
        non_empty_columns = [sheet.cell(row=1, column=col[0].column).value for col in sheet.columns if any(cell.value is not None for cell in col)]
        return non_empty_rows, non_empty_columns

    def extract_data_to_json(self, selected_sheet, category):
        json_objects = {}
        header = [cell.value.lower() for cell in selected_sheet[1] if cell.value]
        
        if 'code' not in header or 'label' not in header:
            print("Error: 'code' or 'label' column not found in the sheet.")
            return None, None

        code_idx, label_idx = header.index('code'), header.index('label')
        uid_counts = {}
        
        # Check if 'cpl' column exists in header
        cpl_exists = 'cpl' in header
        
        for row in selected_sheet.iter_rows(min_row=2, values_only=True):
            code, label = row[code_idx], row[label_idx]
            if code is not None and label is not None:
                uid = self.generate_uid(code, uid_counts)
                self.uids.append(uid)
                json_object = {
                    "code": code,
                    "category": category,
                    "label": label,
                    "properties": {header[i]: row[i] for i in range(len(header)) if row[i] is not None and i not in {code_idx, label_idx}},
                    "restrictions": {}
                }
                
                # Add 'cpl' to 'restrictions' if it exists in header
                if cpl_exists:
                    cpl_idx = header.index('cpl')
                    json_object["restrictions"]["cpl"] = bool(row[cpl_idx])
                    
                    # Remove 'cpl' from 'properties' if it was added mistakenly
                    if 'cpl' in json_object["properties"]:
                        del json_object["properties"]["cpl"]

                json_objects[uid] = json_object

        if json_objects:
            base_file_name = os.path.splitext(os.path.basename(self.file_path))[0].replace(" ", "_").lower()
            tab_name = selected_sheet.title.replace(" ", "_").lower()
            json_file_name = f"{base_file_name}_{tab_name}.json"
            JSONFileManager.save_json_to_file(json_objects, json_file_name)
            return json_file_name, json_objects
        else:
            return None, None

    @staticmethod
    def generate_uid(code, uid_counts):
        formatted_code = re.sub(r'[^a-zA-Z0-9]', '', code).lower()
        uid_counts[formatted_code] = uid_counts.get(formatted_code, 0) + 1
        uid_suffix = f"d{uid_counts[formatted_code]}" if uid_counts[formatted_code] > 1 else ""
        return formatted_code + uid_suffix

    @staticmethod
    def format_codes_in_column(sheet, column_index):
        formatted_codes = []
        for cell in sheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index, values_only=True):
            code = cell[0]
            if code:
                # Format the code to remove spaces, symbols, and ".pdf"
                formatted_code = re.sub(r'[^a-zA-Z0-9]', '', code).lower()
                formatted_code = formatted_code.replace('pdf', '')
                formatted_codes.append(formatted_code)
        return formatted_codes

    @staticmethod
    def extract_am_best_to_json(selected_sheet):
        header = [cell.value.lower() for cell in selected_sheet[1] if cell.value]

        # Convert column names to lowercase for case-insensitive comparison
        lowercase_header = [col.lower() for col in header]

        required_columns = ['am best #', 'business name', 'naic / aiin', 'rating', 'size category', 'original business name']
        missing_columns = [col for col in required_columns if col.lower() not in lowercase_header]

        if missing_columns:
            print(f"Error: Required columns not found in the sheet: {', '.join(missing_columns)}")
            return

        am_best_idx = lowercase_header.index('am best #')
        business_name_idx = lowercase_header.index('business name')
        naic_idx = lowercase_header.index('naic / aiin')
        rating_idx = lowercase_header.index('rating')
        size_category_idx = lowercase_header.index('size category')
        original_business_name_idx = lowercase_header.index('original business name')

        json_objects = {}

        for row in selected_sheet.iter_rows(min_row=2, values_only=True):
            am_best = row[am_best_idx]
            if am_best is not None:
                am_best = int(am_best)
                json_object = {
                    "am": am_best,
                    "name": row[business_name_idx],
                    "naic": row[naic_idx],
                    "rating": row[rating_idx],
                    "size": row[size_category_idx],
                    "business": row[original_business_name_idx]
                }
                json_objects[am_best] = json_object

        if json_objects:
            with open("index_amb.json", "w") as json_file:
                json.dump(json_objects, json_file, indent=2)
            print(f"\nSuccessfully created {len(json_objects)} JSON objects based on AM Best data.")
        else:
            print("No AM Best data found in the selected sheet.")

        return ExcelDataExtractor

class JSONFileManager:
    @staticmethod
    def save_json_to_file(json_data, file_name):
        with open(file_name, "w") as json_file:
            json.dump(json_data, json_file, indent=2)
        print(f"\nSuccessfully created {len(json_data)} JSON objects.")


class MenuHandler:
    @staticmethod
    def display_main_menu():
        print("\nMain Menu:\n1. Tab Reader\n0. Quit")

    @staticmethod
    def display_tab_menu(sheet_names):
        print("\nAvailable Tabs:\n" + "\n".join([f"{i+1}. {sheet_name}" for i, sheet_name in enumerate(sheet_names)]))

    @staticmethod
    def display_tab_options(tab_name):
        print(f"\nTab Menu:\n1. View Cell Value\n2. Extract and Convert to JSON\n3. Compare with CPL\n4. AM Best to JSON\n0. Back to Main Menu")
        return input("Enter your choice: ")

    @staticmethod
    def display_column_names(sheet):
        print("Column Names:")
        for i, col_name in enumerate(sheet[1], 1):
            print(f"{i}. {col_name.value}")

    @staticmethod
    def extract_am_best_to_json(excel_data_extractor, selected_sheet):
        ExcelDataExtractor.extract_am_best_to_json(selected_sheet)

    @staticmethod
    def compare_with_cpl(excel_data_extractor, selected_sheet, json_file_name):
        MenuHandler.display_tab_menu(excel_data_extractor.workbook.sheetnames)
        comparison_tab_choice = int(input("Select a tab for comparison (enter the corresponding number): ")) - 1
        try:
            comparison_sheet = excel_data_extractor.get_selected_sheet(comparison_tab_choice)
            non_empty_rows, non_empty_columns = excel_data_extractor.count_non_empty_rows_and_columns(comparison_sheet)
            print(f"\nNumber of non-empty rows in '{comparison_sheet.title}': {non_empty_rows}")
            print(f"Number of non-empty columns in '{comparison_sheet.title}': {len(non_empty_columns)}")
            MenuHandler.display_column_names(comparison_sheet)
            
            # Get column index where codes are stored
            column_index = int(input("Enter the column index where codes are stored (enter the corresponding number): ")) - 1
            value_column_index = int(input("Enter the column index where values are stored (enter the corresponding number): ")) - 1
            

            # Format codes from the comparison sheet
            formatted_codes = ExcelDataExtractor.format_codes_in_column(comparison_sheet, column_index)
            formatted_values = []
            for cell in comparison_sheet.iter_rows(min_row=2, min_col=value_column_index+1, max_col=value_column_index+1, values_only=True):
                value = cell[0]
                if value:
                    formatted_values.append(value.lower())

            # Compare codes and update JSON data
            with open(json_file_name, "r") as json_file:
                json_data = json.load(json_file)

            changes_count = 0
            for code, value in zip(formatted_codes, formatted_values):
                if code in json_data:
                    if value == "unacceptable":
                        json_data[code]["restrictions"]["cpl"] = True
                        changes_count += 1
                    else:
                        json_data[code]["restrictions"]["cpl"] = False
                        changes_count += 1


            # Save updated JSON data back to file
            with open(json_file_name, "w") as json_file:
                json.dump(json_data, json_file, indent=2)

            print(f"\n{changes_count} JSON objects updated with CPL information.")

        except IndexError:
            print("Invalid tab selection for comparison. Please try again.")


def main():
    excel_data_extractor = ExcelDataExtractor('index.xlsx')
    json_file_name = None

    while True:
        MenuHandler.display_main_menu()
        choice = input("Enter your choice: ")

        if choice == '1':
            MenuHandler.display_tab_menu(excel_data_extractor.workbook.sheetnames)
            tab_choice = int(input("Select a tab (enter the corresponding number): ")) - 1

            try:
                selected_sheet = excel_data_extractor.get_selected_sheet(tab_choice)
                non_empty_rows, non_empty_columns = excel_data_extractor.count_non_empty_rows_and_columns(selected_sheet)
                print(f"\nNumber of non-empty rows in '{selected_sheet.title}': {non_empty_rows}")
                print(f"Number of non-empty columns in '{selected_sheet.title}': {len(non_empty_columns)}")
                MenuHandler.display_column_names(selected_sheet)

                while True:
                    choice = MenuHandler.display_tab_options(selected_sheet.title)
                    if choice == '1':
                        cell_address = input("Enter the cell address (e.g., A1): ")
                        try:
                            print(f"Value at {cell_address}: {selected_sheet[cell_address].value}")
                        except Exception as e:
                            print(f"Error: {e}")
                    elif choice == '2':
                        category = input("Enter the category for extraction: ")
                        json_file_name, json_objects = excel_data_extractor.extract_data_to_json(selected_sheet, category)
                        if json_objects is not None:
                            print(f"Generated JSON file: {json_file_name}")
                    elif choice == '3':
                        if json_file_name:
                            MenuHandler.compare_with_cpl(excel_data_extractor, selected_sheet, json_file_name)
                        else:
                            print("No JSON file generated yet. Please choose option 2 first.")
                    elif choice == '4':
                        if selected_sheet:
                            MenuHandler.extract_am_best_to_json(excel_data_extractor, selected_sheet)
                        else:
                            print("Failed to target the AM Best index.")
                    elif choice == '0':
                        break
                    else:
                        print("Please enter a valid option for the Tab Menu.")

            except IndexError:
                print("Invalid tab selection. Please try again.")

        elif choice == '0':
            break

        else:
            print("Please enter a valid option for the Main Menu.")

if __name__ == "__main__":
    main()
    input("Press any key to exit...")
